from fastapi import APIRouter, Depends, Request, Form
from fastapi.responses import RedirectResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date

from app.base_datos import get_db
from app.modelos.modelos import Usuario, Espacio, Reserva

router = APIRouter()

def verificar_admin(request: Request):
    return request.cookies.get("usuario_rol") == "admin"

@router.get("/dashboard")
async def admin_dashboard(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    espacios = db.query(Espacio).all()
    total_espacios = len(espacios)
    disponibles = sum(1 for e in espacios if e.estado == "disponible")
    ocupados = sum(1 for e in espacios if e.estado == "ocupado")
    reservados = sum(1 for e in espacios if e.estado == "reservado")
    en_mantenimiento = sum(1 for e in espacios if e.estado == "mantenimiento")
    porcentaje_ocupacion = round((ocupados + reservados) / total_espacios * 100) if total_espacios > 0 else 0

    total_usuarios = db.query(Usuario).count()
    reservas_activas = db.query(Reserva).filter(Reserva.estado == "activa").count()

    todas_reservas = db.query(Reserva).all()
    hoy = date.today()
    reservas_hoy = sum(1 for r in todas_reservas if r.fecha_reserva and r.fecha_reserva.date() == hoy)

    ingresos_hoy = db.query(func.sum(Reserva.monto_pagado)).filter(
        Reserva.estado == "activa"
    ).scalar() or 0

    ultimas_reservas = db.query(Reserva).order_by(Reserva.fecha_reserva.desc()).limit(8).all()
    for r in ultimas_reservas:
        r.espacio = db.query(Espacio).filter(Espacio.id_espacio == r.id_espacio).first()
        r.usuario = db.query(Usuario).filter(Usuario.id_usuario == r.id_usuario).first()

    pisos = {}
    for e in espacios:
        pisos.setdefault(e.piso, []).append(e)

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_dashboard.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "total_espacios": total_espacios,
            "disponibles": disponibles,
            "ocupados": ocupados,
            "reservados": reservados,
            "en_mantenimiento": en_mantenimiento,
            "porcentaje_ocupacion": porcentaje_ocupacion,
            "total_usuarios": total_usuarios,
            "reservas_activas": reservas_activas,
            "reservas_hoy": reservas_hoy,
            "ingresos_hoy": ingresos_hoy,
            "ultimas_reservas": ultimas_reservas,
            "pisos": pisos,
        }
    )


# ====================== GESTIÓN DE ESPACIOS ======================

@router.get("/espacios")
async def listar_espacios(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    espacios = db.query(Espacio).order_by(Espacio.piso, Espacio.numero_espacio).all()
    total = len(espacios)
    disponibles = sum(1 for e in espacios if e.estado == "disponible")
    ocupados = sum(1 for e in espacios if e.estado == "ocupado")
    reservados = sum(1 for e in espacios if e.estado == "reservado")
    mantenimiento = sum(1 for e in espacios if e.estado == "mantenimiento")

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_espacios.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "espacios": espacios,
            "total": total,
            "disponibles": disponibles,
            "ocupados": ocupados,
            "reservados": reservados,
            "mantenimiento": mantenimiento,
            "exito": request.query_params.get("exito"),
            "error": request.query_params.get("error"),
        }
    )


@router.post("/espacios/crear")
async def crear_espacio(
    request: Request,
    numero_espacio: str = Form(...),
    piso: int = Form(...),
    tipo_espacio: str = Form(...),
    estado: str = Form("disponible"),
    tarifa_por_hora: float = Form(...),
    tarifa_por_dia: float = Form(...),
    tarifa_mensual: float = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    existente = db.query(Espacio).filter(Espacio.numero_espacio == numero_espacio).first()
    if existente:
        return RedirectResponse(url="/admin/espacios?error=El+número+de+espacio+ya+existe", status_code=303)

    nuevo = Espacio(
        numero_espacio=numero_espacio.upper(),
        piso=piso,
        tipo_espacio=tipo_espacio,
        estado=estado,
        tarifa_por_hora=tarifa_por_hora,
        tarifa_por_dia=tarifa_por_dia,
        tarifa_mensual=tarifa_mensual,
    )
    db.add(nuevo)
    db.commit()
    return RedirectResponse(url="/admin/espacios?exito=Espacio+creado+correctamente", status_code=303)


@router.post("/espacios/editar/{espacio_id}")
async def editar_espacio(
    request: Request,
    espacio_id: int,
    numero_espacio: str = Form(...),
    piso: int = Form(...),
    tipo_espacio: str = Form(...),
    estado: str = Form(...),
    tarifa_por_hora: float = Form(...),
    tarifa_por_dia: float = Form(...),
    tarifa_mensual: float = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    espacio = db.query(Espacio).filter(Espacio.id_espacio == espacio_id).first()
    if espacio:
        espacio.numero_espacio = numero_espacio.upper()
        espacio.piso = piso
        espacio.tipo_espacio = tipo_espacio
        espacio.estado = estado
        espacio.tarifa_por_hora = tarifa_por_hora
        espacio.tarifa_por_dia = tarifa_por_dia
        espacio.tarifa_mensual = tarifa_mensual
        db.commit()
    return RedirectResponse(url="/admin/espacios?exito=Espacio+actualizado+correctamente", status_code=303)


@router.post("/espacios/eliminar/{espacio_id}")
async def eliminar_espacio(request: Request, espacio_id: int, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    espacio = db.query(Espacio).filter(Espacio.id_espacio == espacio_id).first()
    if espacio:
        db.delete(espacio)
        db.commit()
    return RedirectResponse(url="/admin/espacios?exito=Espacio+eliminado+correctamente", status_code=303)


@router.post("/espacios/estado/{espacio_id}")
async def cambiar_estado_espacio(
    request: Request,
    espacio_id: int,
    estado: str = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    espacio = db.query(Espacio).filter(Espacio.id_espacio == espacio_id).first()
    if espacio:
        espacio.estado = estado
        db.commit()
    return RedirectResponse(url="/admin/espacios?exito=Estado+actualizado", status_code=303)

# ====================== GESTIÓN DE USUARIOS ======================

@router.get("/usuarios")
async def listar_usuarios(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    usuarios = db.query(Usuario).order_by(Usuario.fecha_registro.desc()).all()
    total = len(usuarios)
    activos = sum(1 for u in usuarios if u.estado == "activo")
    inactivos = sum(1 for u in usuarios if u.estado != "activo")
    admins = sum(1 for u in usuarios if u.tipo_usuario == "admin")

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_usuarios.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "usuarios": usuarios,
            "total": total,
            "activos": activos,
            "inactivos": inactivos,
            "admins": admins,
            "exito": request.query_params.get("exito"),
            "error": request.query_params.get("error"),
        }
    )


@router.post("/usuarios/editar/{usuario_id}")
async def editar_usuario(
    request: Request,
    usuario_id: int,
    nombre: str = Form(...),
    apellido: str = Form(...),
    email: str = Form(...),
    dni: str = Form(...),
    telefono: str = Form(...),
    direccion: str = Form(""),
    tipo_usuario: str = Form(...),
    estado: str = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    usuario = db.query(Usuario).filter(Usuario.id_usuario == usuario_id).first()
    if usuario:
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.email = email
        usuario.dni = dni
        usuario.telefono = telefono
        usuario.direccion = direccion
        usuario.tipo_usuario = tipo_usuario
        usuario.estado = estado
        db.commit()
    return RedirectResponse(url="/admin/usuarios?exito=Usuario+actualizado+correctamente", status_code=303)


@router.post("/usuarios/toggle/{usuario_id}")
async def toggle_usuario(request: Request, usuario_id: int, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    usuario = db.query(Usuario).filter(Usuario.id_usuario == usuario_id).first()
    if usuario:
        usuario.estado = "inactivo" if usuario.estado == "activo" else "activo"
        db.commit()
    return RedirectResponse(url="/admin/usuarios?exito=Estado+de+usuario+actualizado", status_code=303)


@router.post("/usuarios/reset-password/{usuario_id}")
async def reset_password(
    request: Request,
    usuario_id: int,
    nueva_password: str = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

    usuario = db.query(Usuario).filter(Usuario.id_usuario == usuario_id).first()
    if usuario:
        usuario.contrasena = pwd_context.hash(nueva_password)
        db.commit()
    return RedirectResponse(url="/admin/usuarios?exito=Contraseña+restablecida+correctamente", status_code=303)

# ====================== GESTIÓN DE RESERVAS ======================

from app.crud import reservas_admin as crud_reservas
from typing import Optional
from sqlalchemy import func

ESTADOS_RESERVA = ["activa", "completada", "cancelada", "pendiente"]

@router.get("/reservas")
async def listar_reservas(
    request: Request,
    db: Session = Depends(get_db),
    estado: Optional[str] = None,
    fecha: Optional[str] = None,
    buscar: Optional[str] = None,
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    reservas     = crud_reservas.obtener_todas(db, estado, fecha, buscar)
    estadisticas = crud_reservas.stats(db)

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_reservas.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "reservas":      reservas,
            "estadisticas":  estadisticas,
            "estados":       ESTADOS_RESERVA,
            "filtro_estado": estado or "todos",
            "filtro_fecha":  fecha  or "",
            "filtro_buscar": buscar or "",
            "exito": request.query_params.get("exito"),
            "error": request.query_params.get("error"),
        }
    )


@router.get("/reservas/{id_reserva}")
async def detalle_reserva(request: Request, id_reserva: int, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    reserva = crud_reservas.obtener_por_id(db, id_reserva)
    if not reserva:
        return RedirectResponse(url="/admin/reservas?error=Reserva+no+encontrada", status_code=303)

    otras = crud_reservas.otras_reservas_usuario(db, reserva.id_usuario, id_reserva)

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_reserva_detalle.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "reserva":        reserva,
            "otras_reservas": otras,
            "estados":        ESTADOS_RESERVA,
            "exito": request.query_params.get("exito"),
            "error": request.query_params.get("error"),
        }
    )


@router.post("/reservas/{id_reserva}/estado")
async def cambiar_estado_reserva(request: Request, id_reserva: int, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    form   = await request.form()
    estado = form.get("estado")
    crud_reservas.cambiar_estado(db, id_reserva, estado)
    return RedirectResponse(url=f"/admin/reservas/{id_reserva}?exito=Estado+actualizado", status_code=303)


@router.post("/reservas/{id_reserva}/cancelar")
async def cancelar_reserva(request: Request, id_reserva: int, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    ok = crud_reservas.cancelar(db, id_reserva)
    if not ok:
        return RedirectResponse(url=f"/admin/reservas/{id_reserva}?error=No+se+pudo+cancelar", status_code=303)
    return RedirectResponse(url=f"/admin/reservas/{id_reserva}?exito=Reserva+cancelada", status_code=303)

# ====================== GESTIÓN DE PERSONAL ======================

@router.get("/personal")
async def listar_personal(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Personal, Puesto
    personal = db.query(Personal).all()

    for p in personal:
        p.usuario = db.query(Usuario).filter(Usuario.id_usuario == p.id_usuario).first()
        p.puesto = db.query(Puesto).filter(Puesto.id_puesto == p.id_puesto).first()

    puestos = db.query(Puesto).all()

    # Solo usuarios con rol personal sin registro en tabla personal
    usuarios_personal = db.query(Usuario).filter(
        Usuario.tipo_usuario == "personal"
    ).all()

    total = len(personal)
    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_personal.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "personal": personal,
            "puestos": puestos,
            "usuarios_personal": usuarios_personal,
            "total": total,
            "exito": request.query_params.get("exito"),
            "error": request.query_params.get("error"),
        }
    )


@router.post("/personal/crear")
async def crear_personal(
    request: Request,
    id_usuario: int = Form(...),
    id_puesto: int = Form(...),
    turno: str = Form(...),
    fecha_contratacion: str = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Personal
    from datetime import date

    existente = db.query(Personal).filter(Personal.id_usuario == id_usuario).first()
    if existente:
        return RedirectResponse(url="/admin/personal?error=Este+usuario+ya+tiene+registro+de+personal", status_code=303)

    nuevo = Personal(
        id_usuario=id_usuario,
        id_puesto=id_puesto,
        turno=turno,
        fecha_contratacion=date.fromisoformat(fecha_contratacion)
    )
    db.add(nuevo)
    db.commit()
    return RedirectResponse(url="/admin/personal?exito=Personal+registrado+correctamente", status_code=303)


@router.post("/personal/editar/{personal_id}")
async def editar_personal(
    request: Request,
    personal_id: int,
    id_puesto: int = Form(...),
    turno: str = Form(...),
    fecha_contratacion: str = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Personal
    from datetime import date

    p = db.query(Personal).filter(Personal.id_personal == personal_id).first()
    if p:
        p.id_puesto = id_puesto
        p.turno = turno
        p.fecha_contratacion = date.fromisoformat(fecha_contratacion)
        db.commit()
    return RedirectResponse(url="/admin/personal?exito=Personal+actualizado+correctamente", status_code=303)


@router.post("/personal/eliminar/{personal_id}")
async def eliminar_personal(request: Request, personal_id: int, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Personal
    p = db.query(Personal).filter(Personal.id_personal == personal_id).first()
    if p:
        db.delete(p)
        db.commit()
    return RedirectResponse(url="/admin/personal?exito=Personal+eliminado+correctamente", status_code=303)


@router.post("/puestos/crear")
async def crear_puesto(
    request: Request,
    nombre_puesto: str = Form(...),
    salario_base: float = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Puesto
    nuevo = Puesto(nombre_puesto=nombre_puesto, salario_base=salario_base)
    db.add(nuevo)
    db.commit()
    return RedirectResponse(url="/admin/personal?exito=Puesto+creado+correctamente", status_code=303)

# ====================== GESTIÓN DE PAGOS ======================

@router.get("/pagos")
async def listar_pagos(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Pago
    pagos = db.query(Pago).order_by(Pago.fecha_pago.desc()).all()

    for p in pagos:
        p.usuario = db.query(Usuario).filter(Usuario.id_usuario == p.id_usuario).first()

    total_pagos = len(pagos)
    total_ingresos = sum(float(p.monto or 0) for p in pagos if p.estado == "pagado")
    pagos_pendientes = sum(1 for p in pagos if p.estado == "pendiente")
    pagos_completados = sum(1 for p in pagos if p.estado == "pagado")

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_pagos.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "pagos": pagos,
            "total_pagos": total_pagos,
            "total_ingresos": total_ingresos,
            "pagos_pendientes": pagos_pendientes,
            "pagos_completados": pagos_completados,
            "exito": request.query_params.get("exito"),
        }
    )


@router.post("/pagos/registrar")
async def registrar_pago_manual(
    request: Request,
    id_usuario: int = Form(...),
    monto: float = Form(...),
    metodo_pago: str = Form(...),
    comprobante: str = Form(""),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Pago
    from datetime import datetime

    nuevo_pago = Pago(
        id_usuario=id_usuario,
        monto=monto,
        metodo_pago=metodo_pago,
        comprobante=comprobante,
        estado="pagado",
        fecha_pago=datetime.now()
    )
    db.add(nuevo_pago)
    db.commit()
    return RedirectResponse(url="/admin/pagos?exito=Pago+registrado+correctamente", status_code=303)


@router.post("/pagos/actualizar/{pago_id}")
async def actualizar_estado_pago(
    request: Request,
    pago_id: int,
    estado: str = Form(...),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Pago
    pago = db.query(Pago).filter(Pago.id_pago == pago_id).first()
    if pago:
        pago.estado = estado
        db.commit()
    return RedirectResponse(url="/admin/pagos?exito=Estado+de+pago+actualizado", status_code=303)

# ====================== MANTENIMIENTO ======================

@router.get("/mantenimiento")
async def listar_mantenimiento(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Mantenimiento, Personal

    registros = db.query(Mantenimiento).order_by(Mantenimiento.fecha_inicio.desc()).all()

    for r in registros:
        r.espacio = db.query(Espacio).filter(Espacio.id_espacio == r.id_espacio).first()
        r.personal_obj = db.query(Personal).filter(Personal.id_personal == r.id_personal).first()
        if r.personal_obj:
            r.personal_obj.usuario = db.query(Usuario).filter(
                Usuario.id_usuario == r.personal_obj.id_usuario
            ).first()

    espacios = db.query(Espacio).order_by(Espacio.piso, Espacio.numero_espacio).all()
    personal_list = db.query(Personal).all()
    for p in personal_list:
        p.usuario = db.query(Usuario).filter(Usuario.id_usuario == p.id_usuario).first()

    total = len(registros)
    activos = sum(1 for r in registros if r.estado == "en_proceso")
    completados = sum(1 for r in registros if r.estado == "completado")
    costo_total = sum(float(r.costo or 0) for r in registros)

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_mantenimiento.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "registros": registros,
            "espacios": espacios,
            "personal_list": personal_list,
            "total": total,
            "activos": activos,
            "completados": completados,
            "costo_total": costo_total,
            "exito": request.query_params.get("exito"),
            "error": request.query_params.get("error"),
        }
    )


@router.post("/mantenimiento/crear")
async def crear_mantenimiento(
    request: Request,
    id_espacio: int = Form(...),
    id_personal: int = Form(...),
    tipo_mantenimiento: str = Form(...),
    descripcion: str = Form(""),
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(""),
    costo: float = Form(0),
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Mantenimiento
    from datetime import datetime

    nuevo = Mantenimiento(
        id_espacio=id_espacio,
        id_personal=id_personal,
        tipo_mantenimiento=tipo_mantenimiento,
        descripcion=descripcion,
        fecha_inicio=datetime.fromisoformat(fecha_inicio),
        fecha_fin=datetime.fromisoformat(fecha_fin) if fecha_fin else None,
        estado="en_proceso",
        costo=costo
    )
    db.add(nuevo)

    # Cambiar estado del espacio a mantenimiento
    espacio = db.query(Espacio).filter(Espacio.id_espacio == id_espacio).first()
    if espacio:
        espacio.estado = "mantenimiento"

    db.commit()
    return RedirectResponse(url="/admin/mantenimiento?exito=Mantenimiento+registrado+correctamente", status_code=303)


@router.post("/mantenimiento/completar/{mantenimiento_id}")
async def completar_mantenimiento(
    request: Request,
    mantenimiento_id: int,
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Mantenimiento
    from datetime import datetime

    registro = db.query(Mantenimiento).filter(
        Mantenimiento.id_mantenimiento == mantenimiento_id
    ).first()

    if registro:
        registro.estado = "completado"
        registro.fecha_fin = datetime.now()

        # Liberar el espacio
        espacio = db.query(Espacio).filter(Espacio.id_espacio == registro.id_espacio).first()
        if espacio:
            espacio.estado = "disponible"

        db.commit()

    return RedirectResponse(url="/admin/mantenimiento?exito=Mantenimiento+completado+espacio+liberado", status_code=303)


@router.post("/mantenimiento/eliminar/{mantenimiento_id}")
async def eliminar_mantenimiento(
    request: Request,
    mantenimiento_id: int,
    db: Session = Depends(get_db)
):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from app.modelos.modelos import Mantenimiento

    registro = db.query(Mantenimiento).filter(
        Mantenimiento.id_mantenimiento == mantenimiento_id
    ).first()
    if registro:
        db.delete(registro)
        db.commit()

    return RedirectResponse(url="/admin/mantenimiento?exito=Registro+eliminado+correctamente", status_code=303)

# ====================== REPORTES ======================

@router.get("/reportes")
async def ver_reportes(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    from datetime import datetime, timedelta

    # Reservas por estado
    total_reservas = db.query(Reserva).count()
    reservas_activas = db.query(Reserva).filter(Reserva.estado == "activa").count()
    reservas_canceladas = db.query(Reserva).filter(Reserva.estado == "cancelada").count()

    # Ingresos totales
    ingresos_totales = db.query(func.sum(Reserva.monto_pagado)).filter(
        Reserva.estado == "activa"
    ).scalar() or 0

    # Espacios
    espacios = db.query(Espacio).all()
    total_espacios = len(espacios)
    disponibles = sum(1 for e in espacios if e.estado == "disponible")
    ocupados = sum(1 for e in espacios if e.estado == "ocupado")

    # Usuarios
    total_usuarios = db.query(Usuario).count()
    usuarios_activos = db.query(Usuario).filter(Usuario.estado == "activo").count()

    # Últimas 10 reservas
    ultimas_reservas = db.query(Reserva).order_by(
        Reserva.fecha_reserva.desc()
    ).limit(10).all()
    for r in ultimas_reservas:
        r.espacio = db.query(Espacio).filter(Espacio.id_espacio == r.id_espacio).first()
        r.usuario = db.query(Usuario).filter(Usuario.id_usuario == r.id_usuario).first()

    # Espacios más usados
    from sqlalchemy import desc
    espacios_usados = db.query(
        Reserva.id_espacio,
        func.count(Reserva.id_reserva).label("total")
    ).group_by(Reserva.id_espacio).order_by(desc("total")).limit(5).all()

    espacios_top = []
    for eu in espacios_usados:
        esp = db.query(Espacio).filter(Espacio.id_espacio == eu.id_espacio).first()
        if esp:
            espacios_top.append({"espacio": esp, "total": eu.total})

    return request.app.state.plantillas.TemplateResponse(
        request=request,
        name="admin_reportes.html",
        context={
            "usuario_nombre": request.cookies.get("usuario_nombre", "Admin"),
            "total_reservas": total_reservas,
            "reservas_activas": reservas_activas,
            "reservas_canceladas": reservas_canceladas,
            "ingresos_totales": ingresos_totales,
            "total_espacios": total_espacios,
            "disponibles": disponibles,
            "ocupados": ocupados,
            "total_usuarios": total_usuarios,
            "usuarios_activos": usuarios_activos,
            "ultimas_reservas": ultimas_reservas,
            "espacios_top": espacios_top,
        }
    )


@router.get("/reportes/exportar/excel")
async def exportar_excel(request: Request, db: Session = Depends(get_db)):
    if not verificar_admin(request):
        return RedirectResponse(url="/auth/login", status_code=303)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io
    from datetime import datetime

    wb = openpyxl.Workbook()

    def ajustar_anchos(ws):
        for col in ws.columns:
            col_letter = None
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                    break
            if col_letter:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col if hasattr(cell, 'value')),
                    default=0
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # ===== HOJA 1: RESERVAS =====
    ws1 = wb.active
    ws1.title = "Reservas"

    ws1.merge_cells("A1:I1")
    ws1["A1"] = "REPORTE DE RESERVAS - SMART PARKING"
    ws1["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="0EA5E9")
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, size=10)

    headers = ["#", "Usuario", "Email", "Espacio", "Piso", "Inicio", "Fin", "Monto", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    reservas = db.query(Reserva).order_by(Reserva.fecha_reserva.desc()).all()
    for i, r in enumerate(reservas, 1):
        usuario = db.query(Usuario).filter(Usuario.id_usuario == r.id_usuario).first()
        espacio = db.query(Espacio).filter(Espacio.id_espacio == r.id_espacio).first()
        row = [
            i,
            f"{usuario.nombre} {usuario.apellido}" if usuario else "—",
            usuario.email if usuario else "—",
            espacio.numero_espacio if espacio else "—",
            f"Piso {espacio.piso}" if espacio else "—",
            r.fecha_inicio.strftime("%d/%m/%Y %H:%M") if r.fecha_inicio else "—",
            r.fecha_fin.strftime("%d/%m/%Y %H:%M") if r.fecha_fin else "—",
            float(r.monto_pagado) if r.monto_pagado else 0,
            r.estado or "—",
        ]
        for col, val in enumerate(row, 1):
            ws1.cell(row=i + 3, column=col, value=val)

    ajustar_anchos(ws1)

    # ===== HOJA 2: ESPACIOS =====
    ws2 = wb.create_sheet("Espacios")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "REPORTE DE ESPACIOS - SMART PARKING"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="0EA5E9")
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A2"].font = Font(italic=True, size=10)

    headers2 = ["#", "Número", "Piso", "Tipo", "Estado", "Tarifa/Hora", "Tarifa/Día"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    espacios = db.query(Espacio).order_by(Espacio.piso, Espacio.numero_espacio).all()
    for i, e in enumerate(espacios, 1):
        row = [
            i,
            e.numero_espacio,
            e.piso,
            e.tipo_espacio or "—",
            e.estado or "—",
            float(e.tarifa_por_hora or 0),
            float(e.tarifa_por_dia or 0)
        ]
        for col, val in enumerate(row, 1):
            ws2.cell(row=i + 3, column=col, value=val)

    ajustar_anchos(ws2)

    # ===== HOJA 3: USUARIOS =====
    ws3 = wb.create_sheet("Usuarios")

    ws3.merge_cells("A1:I1")
    ws3["A1"] = "REPORTE DE USUARIOS - SMART PARKING"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="0EA5E9")
    ws3["A1"].alignment = Alignment(horizontal="center")

    ws3["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws3["A2"].font = Font(italic=True, size=10)

    headers3 = ["#", "Nombre", "Apellido", "Email", "DNI", "Teléfono", "Rol", "Estado", "Registro"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    usuarios = db.query(Usuario).order_by(Usuario.fecha_registro.desc()).all()
    for i, u in enumerate(usuarios, 1):
        row = [
            i,
            u.nombre or "—",
            u.apellido or "—",
            u.email or "—",
            u.dni or "—",
            u.telefono or "—",
            u.tipo_usuario or "—",
            u.estado or "—",
            u.fecha_registro.strftime("%d/%m/%Y") if u.fecha_registro else "—"
        ]
        for col, val in enumerate(row, 1):
            ws3.cell(row=i + 3, column=col, value=val)

    ajustar_anchos(ws3)

    # ===== HOJA 4: PAGOS =====
    from app.modelos.modelos import Pago
    ws4 = wb.create_sheet("Pagos")

    ws4.merge_cells("A1:G1")
    ws4["A1"] = "REPORTE DE PAGOS - SMART PARKING"
    ws4["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill = PatternFill("solid", fgColor="0EA5E9")
    ws4["A1"].alignment = Alignment(horizontal="center")

    ws4["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws4["A2"].font = Font(italic=True, size=10)

    headers4 = ["#", "Usuario", "Monto", "Método", "Comprobante", "Fecha", "Estado"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    pagos = db.query(Pago).order_by(Pago.fecha_pago.desc()).all()
    for i, p in enumerate(pagos, 1):
        usuario = db.query(Usuario).filter(Usuario.id_usuario == p.id_usuario).first()
        row = [
            i,
            f"{usuario.nombre} {usuario.apellido}" if usuario else "—",
            float(p.monto or 0),
            p.metodo_pago or "—",
            p.comprobante or "—",
            p.fecha_pago.strftime("%d/%m/%Y %H:%M") if p.fecha_pago else "—",
            p.estado or "—",
        ]
        for col, val in enumerate(row, 1):
            ws4.cell(row=i + 3, column=col, value=val)

    ajustar_anchos(ws4)

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_smart_parking_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )